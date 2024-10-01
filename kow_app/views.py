import base64
import json
import re
from django.shortcuts import redirect, render, get_object_or_404
from .models import *
from django.contrib import auth, messages
from .forms import CategoryForm, ProductForm
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate
from django.db.models import Q
from django.utils.crypto import get_random_string
from django.utils import timezone
from django.template.loader import render_to_string
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.conf import settings



# ---------------- GENERAL SECTION ----------------

from django.shortcuts import render, get_object_or_404
from .models import Category, Products, Cart, CartItem
from django.contrib.auth.decorators import login_required


def index(request):
    categories = Category.objects.all()
    products = Products.objects.all()

    # Similar products logic
    cart_id = _cart_id(request)  # Get the cart_id from session

    # Check if the cart exists
    try:
        cart = Cart.objects.get(cart_id=cart_id)
    except Cart.DoesNotExist:
        cart = None

    similar_products = []
    if cart:
        cart_items = CartItem.objects.filter(cart=cart)
        
        if cart_items.exists():
            # Get the products in the cart
            cart_products = [item.product for item in cart_items]
            
            # Collect unique categories from the cart products
            categories_in_cart = {product.category for product in cart_products}
            
            # Fetch similar products in the same categories but not in the cart
            similar_products = Products.objects.filter(category__in=categories_in_cart).exclude(id__in=[product.id for product in cart_products])

    context = {
        'categories': categories,
        'products': products,
        'similar_products': similar_products,
    }
    return render(request, 'index.html', context)


def all_products(request):
    products = Products.objects.all()
    context = {
        'products': products,
    }
    return render(request, 'all_products.html', context)

def product(request, id):
    product = Products.objects.get(id=id)
    context = {
        'product': product,
    }
    return render(request, 'product.html', context)

def category(request, id):
    category = Category.objects.get(id=id)
    context = {
        'category': category,
    }
    return render(request, 'category.html', context)

# ---------------- GENERAL END ----------------


# --------- AUTHENTICATION AND AUTHORIZATION SECTION ---------
def is_password_strong(password):
    # Password constraint: at least 8 characters, containing at least one uppercase letter, one lowercase letter, one digit, and one special character
    if (
        len(password) < 8
        or not re.search(r"[A-Z]", password)
        or not re.search(r"[a-z]", password)
        or not re.search(r"\d", password)
    ):
        return False
    return True


def login(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']
        user = auth.authenticate(username=email, password=password)
        if user is not None:
            auth.login(request, user)
            if hasattr(user, 'customer'):  # Check if the user is a customer
                messages.success(request, 'Customer LogIn Success!')
                return redirect('index')  # Redirect to customer dashboard
            elif hasattr(user, 'seller'):  # Check if the user is a seller
                messages.success(request, 'Seller LogIn Success!')
                return redirect('s_index')
        else:
            return render(request, 'login.html', {'invalid_credentials':'t'})
    return render(request, 'login.html')

def register(request):
    if request.method == 'POST':
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        email = request.POST['email']
        password = request.POST['password']
        confirm_password = request.POST['confirm_password']
        if password == confirm_password:
            if not is_password_strong(password):
                messages.error(request, 'Passwords Is Too Weak. Use at least 8 characters including uppercase and lowercase letters, digits, and special characters!')
                return render(request, 'register.html')

            if User.objects.filter(email=email).exists():
                print('user exist')
                return render(request, 'register.html', {'username_taken':'p'})
            else:
                user = User.objects.create_user(username=email,
                                                first_name=first_name,
                                                last_name=last_name,
                                                email=email,
                                                password=password)
                customer = Customer.objects.create(user=user, 
                                            first_name=first_name, 
                                            last_name=last_name,
                                            email=email, 
                                            password=password)
                user = authenticate(username=email, password=password)
                auth.login(request, user)
                return redirect('index')
        else:
            return render(request, 'register.html', {'passwords_dont_match':'h'})
    return render(request, 'register.html')

def logout(request):
    auth.logout(request)
    return redirect('/')

# --------- AUTHENTICATION AND AUTHORIZATION END ---------  


# ---------------- USER SECTION ----------------

@login_required
def dashboard(request):
    customer = request.user.customer
    orders = OrderDetail.objects.filter(order__customer=customer)
    context = {
        'customer': customer,
        'orders': orders
    }
    return render(request, "dashboard.html", context)

@login_required
def edit_profile(request):
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        customer = request.user.customer
        customer.first_name = first_name
        customer.last_name = last_name
        customer.email = email
        customer.phone = phone
        customer.address = address
        customer.save()
        messages.success(request, 'Profile updated successfully!')
    return redirect('dashboard')

@login_required
def change_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('oldPassword')
        new_password = request.POST.get('newPassword')

        customer = request.user.customer
        if customer.password == old_password:
            customer.password = new_password 
            customer.save()
            messages.success(request, 'Password changed successfully!')
        else:
            messages.error(request, 'Old password is incorrect. Please try again.')
    return redirect('dashboard')

class CustomTokenGenerator(PasswordResetTokenGenerator):
    def _make_hash_value(self, user, timestamp):
        return (
            str(user.pk) + user.password + str(timestamp)
        )

def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = CustomTokenGenerator().make_token(user)
            reset_password_url = request.build_absolute_uri('/reset_password/{}/{}/'.format(uid, token))
            email_subject = 'Reset Your Password'

            # Render both HTML and plain text versions of the email
            email_body_html = render_to_string('reset_password_email.html', {
                'reset_password_url': reset_password_url,
                'user': user,
            })
            email_body_text = "Click the following link to reset your password: {}".format(reset_password_url)

            # Create an EmailMultiAlternatives object to send both HTML and plain text versions
            email = EmailMultiAlternatives(
                email_subject,
                email_body_text,
                settings.EMAIL_HOST_USER,
                [email],
            )
            email.attach_alternative(email_body_html, 'text/html')  # Attach HTML version
            email.send(fail_silently=False)

            messages.success(request, 'An email has been sent to your email address with instructions on how to reset your password.')
            return redirect('login')
        except User.DoesNotExist:
            messages.error(request, "User with this email does not exist.")
    return render(request, 'forgot_password.html')

def reset_password(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
        customer = Customer.objects.get(user=user)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and CustomTokenGenerator().check_token(user, token):
        if request.method == 'POST':
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')

            if new_password == confirm_password:
                customer.password = new_password
                user.set_password(new_password)
                user.save()
                customer.save()
                messages.success(request, "Password reset successfully. You can now login with your new password.")
                return redirect('login')
            else:
                messages.error(request, "Passwords do not match.")
        return render(request, 'reset_password.html')
    else:
        messages.error(request, "Invalid reset link. Please try again or request a new reset link.")
        return redirect('login')
 
# ---------------- USER END ----------------


# ---------------- CART SECTION ----------------
@login_required(login_url='/login/')
def _cart_id(request):
    cart = request.session.session_key
    if not cart:
        cart = request.session.create()
    return cart


@login_required
def add_cart(request, product_id):
    product = Products.objects.get(id=product_id)
    try:
        cart = Cart.objects.get(cart_id=_cart_id(request))
    except Cart.DoesNotExist:
        cart = Cart.objects.create(cart_id=_cart_id(request))
        cart.save()
    try:
        cart_item = CartItem.objects.get(product=product, cart=cart)
        if cart_item.quantity < cart_item.product.quantity:
            cart_item.quantity += 1
        cart_item.save()
    except CartItem.DoesNotExist:
        cart_item = CartItem.objects.create(product=product, quantity=1, cart=cart)
        cart_item.save()
    return redirect('cart_detail')


@login_required
def cart_detail(request, total=0, counter=0, cart_items=None):
    try:
        customer = request.user.customer
        cart = Cart.objects.get(cart_id=_cart_id(request))
        cart_items = CartItem.objects.filter(cart=cart, active=True)
        for cart_item in cart_items:
            total += (cart_item.product.price * cart_item.quantity)
            counter += cart_item.quantity
    except ObjectDoesNotExist:
        pass
    return render(request, 'cart.html', dict(cart_items=cart_items, total=total, counter=counter, customer=customer, STRIPE_PUBLIC_KEY=settings.STRIPE_PUBLIC_KEY))


@login_required
def cart_remove(request, product_id):
    cart = Cart.objects.get(cart_id=_cart_id(request))
    product = get_object_or_404(Products, id=product_id)
    cart_item = CartItem.objects.get(product=product, cart=cart)
    if cart_item.quantity > 1:
        cart_item.quantity -= 1
        cart_item.save()
    else:
        cart_item.delete()
    return redirect('cart_detail')


@login_required()
def full_remove(request, product_id):
    cart = Cart.objects.get(cart_id=_cart_id(request))
    product = get_object_or_404(Products, id=product_id)
    cart_item = CartItem.objects.get(product=product, cart=cart)
    cart_item.delete()
    return redirect('cart_detail')

# ---------------- CART END ----------------


# ------------------- ORDER SECTION --------------------

def generate_transaction_id(payment_method):
    if payment_method == 'Cash On Delivery':
        timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
        random_string = get_random_string(length=4)
        transaction_id = f'COD-{timestamp}-{random_string}'
        return transaction_id
    else:
        return None

@login_required
def place_order(request):
    try:
        if request.method == 'POST':
            payment_method = request.POST.get('payment')
            address = request.POST.get('address')
            
            if len(address) > 5:
                total = 0
                customer = request.user.customer
                cart = Cart.objects.get(cart_id=_cart_id(request))
                cart_items = CartItem.objects.filter(cart=cart, active=True)
                for cart_item in cart_items:
                    total += (cart_item.product.price * cart_item.quantity)
                order = Order.objects.create(customer=customer, delivery_address=address, total_amount=total)
                
                # payment_method = 'Cash On Delivery'
                transaction_id = generate_transaction_id(payment_method)
                payment = Payment.objects.create(transaction_id=transaction_id, order=order, amount=total, payment_method=payment_method)
                
                for cart_item in cart_items:
                    item = Products.objects.get(pk=cart_item.product.pk)
                    if cart_item.quantity <= item.quantity:
                        OrderDetail.objects.create(order=order, product=cart_item.product, quantity=cart_item.quantity, price=cart_item.product.price)
                        item.quantity -= cart_item.quantity
                        cart_item.product.save()
                cart_items.delete()
                messages.success(request, 'Your Order Has Been Placed!')
            else:
                messages.error(request, 'Invalid Address')
                return redirect('checkout')
        return redirect('dashboard')
    except Exception as e:
        messages.error(request, f'Sorry! Your Order Was Not Placed: {e}')
        return redirect('cart_detail')

@login_required
def order_details(request, pk):
    customer = request.user.customer
    order = OrderDetail.objects.get(pk=pk)
    subtotal = order.price * order.quantity
    context = {
        'customer': customer,
        'order': order,
        'subtotal': subtotal,
    }
    return render(request, "order_details.html", context)

# ------------------- ORDER END --------------------


# ------------------- ADMIN SECTION ---------------------------

def admin_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = auth.authenticate(username=username, password=password)
        if user is not None:
            auth.login(request, user)
            return redirect('home')
        else:
            return render(request, 'admin_login.html', {'invalid_credentials':'t'})
    return render(request, 'admin_login.html')

def home(request):
    products_list = Products.objects.all()
    category_list = Category.objects.all()
    return render(request, 'admin_home.html', {'categories': category_list, 'products': products_list})

def add_category(request):
    if request.method == 'POST':
        name = request.POST['name']
        description = request.POST['description']
        image = request.FILES['image']
        new_category = Category(name=name, description=description, image=image)
        new_category.save()
        return redirect('custom_admin:home')
    return render(request, 'add_category.html')

def add_product(request):
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('home')  
    else:
        form = ProductForm()

    return render(request, 'add_product.html', {'form': form})

def edit_category(request, id):
    category = Category.objects.get(id=id)
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES, instance=category)

        if form.is_valid():
            form.save()
            return redirect('edit_category', id)

    else:
        form = CategoryForm(instance=category)
    return render(request, 'edit_category.html', {'category': category, 'form': form})

def edit_product(request, id):
    product = Products.objects.get(id=id)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)

        if form.is_valid():
            form.save()
            return redirect('edit_product', id)

    else:
        form = ProductForm(instance=product)
    return render(request, 'edit_product.html', {'product': product, 'form': form})

def delete_category(request, id):
    try:
        category = Category.objects.get(id=id)
        category.delete()
        return JsonResponse({'message': 'Category deleted successfully'})
    except Category.DoesNotExist:
        return JsonResponse({'error': 'Category not found'}, status=404)

def delete_product(request, id):
    try:
        product = Products.objects.get(id=id)
        product.delete()
        return JsonResponse({'message': 'Product deleted successfully'})
    except Products.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    
# ------------------- ADMIN END ---------------------------
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.db.models import Sum, F
from django.http import HttpResponse
from django.shortcuts import render
from .models import Order, OrderDetail  # Ensure these models are correctly defined and imported

def monthly_sales_report(request):
    # Get the month from the query parameters, default to the current month and year
    month = request.GET.get('month', datetime.now().strftime('%Y-%m'))
    
    # Split the month and year for query
    year, month = map(int, month.split('-'))

    # Calculate the start and end dates for the selected month
    start_date = datetime(year, month, 1)
    if month < 12:
        end_date = datetime(year, month + 1, 1)
    else:
        end_date = datetime(year + 1, 1, 1)

    # Filter orders within the selected month using the correct field
    orders = Order.objects.filter(order_date__range=[start_date, end_date])
    
    # Calculate total sales for the month
    total_sales = orders.aggregate(total=Sum('total_amount'))['total'] or 0

    # Calculate sales per product for the month
    order_items = OrderDetail.objects.filter(order__in=orders)
    product_sales = order_items.values('product__name').annotate(
        total_quantity=Sum('quantity'),
        total_amount=Sum(F('quantity') * F('product__price'))
    ).order_by('-total_amount')

    if 'export' in request.GET:
        # Create an Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Sales Report {month}-{year}".replace("/", "-")

        # Write the headers
        headers = ['Product', 'Quantity Sold', 'Total Sales Amount']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

        # Write data to the Excel sheet
        for row_num, item in enumerate(product_sales, 2):
            ws[f"A{row_num}"] = item['product__name']
            ws[f"B{row_num}"] = item['total_quantity']
            ws[f"C{row_num}"] = item['total_amount']

        # Add a row for total sales
        ws[f"A{row_num + 1}"] = 'Total Sales'
        ws[f"C{row_num + 1}"] = total_sales

        # Prepare the response to download the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Sales_Report_{month}_{year}.xlsx'
        wb.save(response)
        return response

    # Context for rendering the HTML page
    context = {
        'total_sales': total_sales,
        'month': f"{year}-{month:02d}",
        'product_sales': product_sales,
    }
    return render(request, 'admin_report.html', context)
# ------------------------------------------------------------------------------------------
from django.shortcuts import render, get_object_or_404
from .models import Cart, CartItem, Products
from django.contrib.auth.decorators import login_required

@login_required
def similar_products(request):
    cart_id = _cart_id(request)  # Get the cart_id from session

    # Check if the cart exists
    try:
        cart = Cart.objects.get(cart_id=cart_id)
    except Cart.DoesNotExist:
        return render(request, 'similar_products.html', {'message': "No cart found for this session."})

    # Get all cart items for this cart
    cart_items = CartItem.objects.filter(cart=cart)
    
    if not cart_items.exists():
        return render(request, 'similar_products.html', {'message': "No products found in your cart."})

    # Get the products in the cart
    cart_products = [item.product for item in cart_items]

    # Collect unique categories from the cart products
    categories = {product.category for product in cart_products}

    # Fetch similar products in the same categories but not in the cart
    similar_products = Products.objects.filter(category__in=categories).exclude(id__in=[product.id for product in cart_products])

    # Check if there are any similar products
    message = "No similar products found." if not similar_products.exists() else ""

    context = {
        'cart_products': cart_products,
        'similar_products': similar_products,
        'message': message,
    }

    return render(request, 'similar_products.html', context)

# ---------------------------------------------------------------------------------------------------------------



from django.shortcuts import render
import os
import google.generativeai as genai

# Configure Gemini API with your API key
genai.configure(api_key="AIzaSyBUqosPLlX6kXtBD-5-Fym2fQwYdJUx9Ms")

def fashion_recommendation(request):
    recommendation = None
    
    if request.method == "POST":
        gender = request.POST['gender']
        color = request.POST['color']
        category = request.POST['category']
        climate = request.POST['climate']
        usage = request.POST['usage']
        
        # Define the prompt to send to the Gemini API
        user_input = f"""
        Gender: {gender}
        Colour: {color}
        Category: {category}
        Climate: {climate}
        Usage: {usage}
        """
        
        # Call the Gemini model
        model = genai.GenerativeModel(model_name="gemini-1.5-flash")
        chat_session = model.start_chat(history=[])
        response = chat_session.send_message(user_input)
        
        recommendation = response.text  # Fetch the recommendation result
    
    return render(request, 'fashion_recommendation.html', {'recommendation': recommendation})

# ----------------------------------------------------------------------------------------------------------------

import json
from django.shortcuts import render
from .models import Products

def visualization_view(request):
    # Get selected category from GET parameters
    selected_category = request.GET.get('category', '')

    # Filter products by category if a category is selected
    if selected_category:
        products = Products.objects.filter(category__name=selected_category)
    else:
        products = Products.objects.all()

    categories = products.values_list('category__name', flat=True).distinct()
    product_names = products.values_list('name', flat=True)

    # Prepare dummy data for chart (replace with your actual data processing)
    chart_data = {
        'labels': list(product_names),
        'values': list(range(len(product_names)))  # Dummy data
    }

    context = {
        'chart_data': json.dumps(chart_data),  # Ensure JSON format
        'products': products,
        'categories': categories
    }

    return render(request, 'visualizations.html', context)










import requests
from django.http import JsonResponse
from django.shortcuts import render

# Replace with your API key
API_KEY = 'AIzaSyB64cvdFQmdn9YJnIWyO3pOGB8mKFlPEEE'
API_URL = 'https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent'

# Define predefined responses
PREDEFINED_RESPONSES = {
    "Can you tell me what this THRIVEseeds is about?": "Hi, I’m Cropsy! Our e-commerce website THRIVEseeds specializes in selling high-quality crop seeds for various agricultural needs. We offer a wide range of seeds with detailed descriptions, pricing, and weather-based recommendations.",
    "What kinds of crop seeds do you sell?": "Cropsy here! We offer a diverse range of crop seeds including vegetables, fruits, grains, and pulses. You can browse our categories to find specific types of seeds.",
    "Can you give me details about a specific seed?": "Sure thing! Just tell me the name or category of the seed you’re interested in, and I’ll provide you with more details.",
    "How does weather affect the seeds I should buy?": "Great question! Weather plays a crucial role in crop growth. I can help you choose the right seeds based on your local climate conditions using our weather forecasts.",
    "Can you recommend seeds based on the current weather?": "Absolutely! Based on your location and the current weather conditions, I can suggest the best seeds for optimal growth. Just let me know your location.",
    "How do I add items to my cart?": "To add items to your cart, simply select the desired seed, choose the quantity, and click the 'Add to Cart' button.",
    "I want to remove an item from my cart. How do I do that?": "No problem! Go to your cart page, find the item you want to remove, and click the 'Remove' button next to it.",
    "How do I check out?": "To check out, go to your cart, review the items, and click the 'Proceed to Checkout' button. Follow the prompts to enter your shipping information and payment details.",
    "I have a problem with my order. Who should I contact?": "If you have any issues with your order, please contact our customer support team through the contact form on our website or by email at rdhanya409@gmail.com.",
    "How can I track my order?": "You can track your order by visiting the 'Order Tracking' section on our website and entering your order number.",
    "How do I create an account?": "To create an account, click on the 'Sign Up' button on the homepage, fill out the required information, and submit the form. You’ll receive a confirmation email to complete the registration.",
    "How can I reset my password?": "If you’ve forgotten your password, go to the 'Login' page and click on 'Forgot Password.' Follow the instructions to reset your password.",
    "What weather conditions should I consider when buying seeds?": "When purchasing seeds, you should consider factors like temperature, humidity, rainfall, and soil conditions. I can provide you with weather forecasts to help you make the right decision.",
    "Can you give me today’s weather forecast?": "Sure! Let me check the current weather conditions for your location. Could you share your city or town?",
    "What is the weather forecast for the next 7 days?": "I can provide you with a 7-day weather forecast for your area. Please visit \"weather dashboard\" after login for get weather forecasting data up to 16 days from now.",
    "How does the weather forecasting feature work?": "THRIVEseeds integrates weather data from reliable sources to help you make informed decisions. The forecasts are updated regularly, and I can provide real-time information for your specific area.",
    "What are the available payment options?": "We accept major credit cards, debit cards, UPI, and net banking. You can choose your preferred option during checkout.",
    "How can I contact customer support?": "You can contact our customer support through the contact form on our website or by emailing us at rdhanya409@gmail.com.",
    "What’s your favorite color?": "As much as I’d love to have a favorite color, I’m here to help you with crop seed-related queries! Let me know if you need assistance with any products or weather updates.",
    "Tell me a joke.": "I’m more of a seed and weather expert, but I can certainly help you grow some great crops! Let me know if you need assistance with anything else.",
    "How do I fix my car engine?": "I specialize in crop seeds and weather forecasting, so I might not be able to help with that. However, if you have any questions about our products, I’d be happy to assist!",
    "Can you predict the stock market for me?": "I’m here to provide you with weather forecasts and help with crop seed-related queries. If you’re looking for investment advice, I recommend contacting a financial expert.",
    "How can I grow flowers in space?": "That’s an exciting question! While I can help you grow crops on Earth, space gardening is a bit out of my expertise. Let me know if you need any tips on planting crops here on Earth.",
    "Can you recommend seeds based on the current weather?": "Absolutely! To recommend the best seeds for your area, I need to know your location. Please tell me your city or town.",
    "how are you": "I'm just a chatbot here to assist you with crop seed-related questions. How can I help you today?",
}



def chat(request):
    if request.method == 'POST':
        user_message = request.POST.get('message')

        # Retrieve or initialize conversation history
        conversation_history = request.session.get('conversation_history', [])

        # Add user message to conversation history
        conversation_history.append(f"input: {user_message}")

        # Check if the message matches any predefined response
        bot_reply = PREDEFINED_RESPONSES.get(user_message, None)
        
        if not bot_reply:
            # Define headers and data for the API request
            headers = {
                'Content-Type': 'application/json',
            }
            
            # Prepare context: Use the conversation history
            messages = [{'text': message} for message in conversation_history]
            
            # Prepare data with context (previous conversation)
            data = {
                'contents': [
                    {
                        'parts': messages
                    }
                ]
            }

            # Make the API request
            try:
                response = requests.post(f'{API_URL}?key={API_KEY}', headers=headers, json=data)
                response.raise_for_status()  # Raise an exception for HTTP errors
                
                # Parse the JSON response
                api_response = response.json()
                print("API Response:", api_response)  # For debugging
                
                # Extract the bot reply from the response
                bot_reply = api_response['candidates'][0]['content']['parts'][0]['text']
                
                # Limit the response to a certain number of sentences (e.g., 3)
                bot_reply = '. '.join(bot_reply.split('. ')[:3])  # Limits the response to 3 sentences
                
            except requests.RequestException as e:
                # Handle request errors
                print(f"API request error: {e}")
                bot_reply = 'Sorry, there was an error processing your request.'

        # Add bot response to conversation history
        conversation_history.append(f"output: {bot_reply}")

        # Store updated conversation history in session
        request.session['conversation_history'] = conversation_history

        return JsonResponse({'reply': bot_reply})

    # Render the chat interface if not a POST request
    return render(request, 'chatbot.html')





import io
from django.http import HttpResponse
from django.shortcuts import render
from .models import Products
from openpyxl import Workbook
from django.db.models import F

def stock_report(request):
    # Fetch all products and low stock products
    all_products = Products.objects.all()
    low_stock_products = Products.objects.filter(quantity__lte=F('reorder_level'))
    total_products = all_products.count()
    total_low_stock = low_stock_products.count()

    # Check if the 'export' parameter exists in the GET request
    if request.GET.get('export'):
        # Create a workbook and add a worksheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Stock Report"
        
        # Write the header
        headers = ["Product Name", "Category", "Quantity", "Reorder Level"]
        sheet.append(headers)

        # Write product data rows
        for product in all_products:
            sheet.append([product.name, product.category.name, product.quantity, product.reorder_level])

        # Create an in-memory output file for the workbook
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # Create an HTTP response with the Excel file
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=stock_report.xlsx'
        
        return response

    return render(request, 'stock_report.html', {
        'all_products': all_products,
        'low_stock_products': low_stock_products,
        'total_products': total_products,
        'total_low_stock': total_low_stock
    })


from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Customer, Order

def customer_report(request):
    # Check if the request is for downloading the report
    if request.GET.get('download') == 'true':
        return download_customer_report()

    customers = Customer.objects.all()
    customer_data = []

    for customer in customers:
        orders = Order.objects.filter(customer=customer)
        customer_data.append({
            'customer': customer,
            'order_count': orders.count(),
            'total_spent': sum(order.total_amount for order in orders),
        })

    context = {
        'customers': customer_data,
        'total_customers': customers.count(),
    }

    return render(request, 'customer_report.html', context)

def download_customer_report():
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Report"

    # Write the header row
    ws.append(["Customer Name", "Email", "Phone", "Orders Count", "Total Spent"])

    # Fetch data
    customers = Customer.objects.all()
    for customer in customers:
        orders = Order.objects.filter(customer=customer)
        total_spent = sum(order.total_amount for order in orders)
        ws.append([
            f"{customer.first_name} {customer.last_name}",
            customer.email,
            customer.phone,
            orders.count(),
            total_spent,
        ])

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_report.xlsx'
    wb.save(response)
    return response
import json
from django.shortcuts import render
from .models import Products, Category

def visualization_view(request):
    # Get selected category from GET parameters
    selected_category = request.GET.get('category', '')

    # Filter products by category if a category is selected
    if selected_category:
        products = Products.objects.filter(category__name=selected_category)
    else:
        products = Products.objects.all()

    categories = Category.objects.all()  # Get all categories for the filter
    
    # Extract product names and their corresponding prices
    product_names = products.values_list('name', flat=True)
    product_prices = products.values_list('price', flat=True)

    # Convert prices from Decimal to float for JSON serialization
    chart_data = {
        'labels': list(product_names),
        'values': [float(price) for price in product_prices]  # Convert Decimal to float
    }

    context = {
        'chart_data': json.dumps(chart_data),  # Ensure JSON format
        'products': products,
        'categories': categories,
        'selected_category': selected_category  # Keep track of the selected category
    }

    return render(request, 'visualizations.html', context)


# views.py

from django.shortcuts import render
from .models import Products, Category
from django.db.models import Sum

def stock_chart(request):
    # Retrieve stock levels categorized by product categories
    stock_data = (
        Products.objects.values('category__name')
        .annotate(total_quantity=Sum('quantity'))
    )
    
    categories = [data['category__name'] for data in stock_data]
    quantities = [data['total_quantity'] for data in stock_data]
    
    context = {
        'categories': categories,
        'quantities': quantities,
    }
    
    return render(request, 'stock_chart.html', context)
